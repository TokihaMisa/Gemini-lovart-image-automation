import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional
from xml.etree import ElementTree

import openpyxl
from openpyxl.utils import get_column_letter

from utils import col_letter_to_openpyxl_idx, ensure_output_dir


@dataclass
class ProductRow:
    id: str
    name_cn: str
    language: str
    selling_points: str
    image_paths: List[str]
    reference_images_are_product: bool = False


def _load_sheet(wb: openpyxl.Workbook, sheet_spec):
    if isinstance(sheet_spec, int):
        return wb.worksheets[sheet_spec]
    return wb[sheet_spec]


def _build_dispimg_map(xlsx_path: str, logger) -> Dict[str, str]:
    """Parse xl/cellimages.xml and rels to map DISPIMG name → media path inside zip."""
    mapping: Dict[str, str] = {}
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            # Parse cellimages.xml for name → rId
            if "xl/cellimages.xml" not in zf.namelist():
                return mapping

            cellimages_xml = zf.read("xl/cellimages.xml")
            ns = {
                "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            }
            root = ElementTree.fromstring(cellimages_xml)
            name_to_rid: Dict[str, str] = {}
            for ci in root.findall("etc:cellImage", ns):
                pic = ci.find("xdr:pic", ns)
                if pic is None:
                    continue
                nv_pr = pic.find("xdr:nvPicPr", ns)
                if nv_pr is None:
                    continue
                cNvPr = nv_pr.find("xdr:cNvPr", ns)
                if cNvPr is None:
                    continue
                img_name = cNvPr.get("name", "")
                blip = pic.find(".//a:blip", ns)
                rid = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed") if blip is not None else None
                if img_name and rid:
                    name_to_rid[img_name] = rid

            # Parse cellimages.xml.rels for rId → target path
            rels_xml = zf.read("xl/_rels/cellimages.xml.rels")
            rels_root = ElementTree.fromstring(rels_xml)
            rels_ns = {"ns": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rid_to_path: Dict[str, str] = {}
            for rel in rels_root.findall("ns:Relationship", rels_ns):
                rid = rel.get("Id", "")
                target = rel.get("Target", "")
                if rid and target and target != "NULL":
                    rid_to_path[rid] = target

            # Build final mapping: name → zip path
            for name, rid in name_to_rid.items():
                if rid in rid_to_path:
                    path_in_zip = "xl/" + rid_to_path[rid]
                    mapping[name] = path_in_zip

            logger.info(f"DISPIMG: mapped {len(mapping)} image(s) from cellimages.xml")
    except Exception as e:
        logger.warning(f"DISPIMG: failed to parse cellimages.xml: {e}")
    return mapping


_DISPIMG_RE = re.compile(r'DISPIMG\("([^"]+)"')


def _extract_dispimg_id(cell_value) -> Optional[str]:
    """If the cell contains a DISPIMG formula, return the image ID."""
    if cell_value is None:
        return None
    s = str(cell_value)
    m = _DISPIMG_RE.search(s)
    return m.group(1) if m else None


def parse_reference_images_are_product(value) -> bool:
    """Return True when the Excel flag says reference images are the same product."""
    text = str(value or "").strip().lower()
    return text in {"是", "yes", "y", "true", "1"}


def resolve_image_scan_config(excel_cfg: dict) -> dict:
    """Resolve image column scanning settings to 1-based openpyxl indexes."""
    image_cfg = excel_cfg.get("image_columns", {})
    start_col = col_letter_to_openpyxl_idx(image_cfg["start"])

    if image_cfg.get("end"):
        end_col = col_letter_to_openpyxl_idx(image_cfg["end"])
    elif image_cfg.get("max_columns"):
        end_col = start_col + int(image_cfg["max_columns"]) - 1
    else:
        end_col = None

    empty_streak = int(image_cfg.get("empty_streak", 2))
    if empty_streak < 1:
        empty_streak = 1
    if end_col is not None and end_col < start_col:
        raise ValueError("excel.image_columns.end must be >= start")

    return {
        "start_col": start_col,
        "end_col": end_col,
        "empty_streak": empty_streak,
    }


def read_products(config: dict, logger, limit: int | None = None) -> List[ProductRow]:
    excel_cfg = config["excel"]
    xlsx_path = excel_cfg["path"]
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    # Parse DISPIMG mapping from zip
    dispimg_map = _build_dispimg_map(str(path), logger)

    # Load workbook (no data_only so we get DISPIMG formulas as text)
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = _load_sheet(wb, excel_cfg["sheet"])
    logger.info(f"Loaded sheet: {ws.title} ({ws.max_row} rows)")

    cols = excel_cfg["columns"]
    id_col = col_letter_to_openpyxl_idx(cols["id"])
    name_col = col_letter_to_openpyxl_idx(cols["name_cn"])
    lang_col = col_letter_to_openpyxl_idx(cols["language"])
    sp_col = col_letter_to_openpyxl_idx(cols["selling_points"])
    ref_flag_col_spec = cols.get("reference_images_are_product")
    ref_flag_col = col_letter_to_openpyxl_idx(ref_flag_col_spec) if ref_flag_col_spec else None
    scan_cfg = resolve_image_scan_config(excel_cfg)

    products: List[ProductRow] = []
    skipped = 0

    for row_idx in range(2, ws.max_row + 1):
        product_id = str(ws.cell(row=row_idx, column=id_col).value or "").strip()
        if not product_id:
            continue

        name_cn = str(ws.cell(row=row_idx, column=name_col).value or "").strip()
        language = str(ws.cell(row=row_idx, column=lang_col).value or "").strip()
        selling_points = str(ws.cell(row=row_idx, column=sp_col).value or "").strip()
        reference_images_are_product = (
            parse_reference_images_are_product(ws.cell(row=row_idx, column=ref_flag_col).value)
            if ref_flag_col
            else False
        )

        if not name_cn:
            logger.warning(f"Row {row_idx}: empty product name, skipping")
            skipped += 1
            continue

        out_dir = ensure_output_dir(product_id)
        image_slots: List[str] = []
        seen_ids = set()

        # Scan image columns
        scan_col = scan_cfg["start_col"]
        empty_streak = 0
        with zipfile.ZipFile(str(path), "r") as zf:
            while empty_streak < scan_cfg["empty_streak"]:
                if scan_cfg["end_col"] is not None and scan_col > scan_cfg["end_col"]:
                    break
                cell_val = ws.cell(row=row_idx, column=scan_col).value
                dispimg_id = _extract_dispimg_id(cell_val)

                if dispimg_id and dispimg_id in dispimg_map and dispimg_id not in seen_ids:
                    empty_streak = 0
                    seen_ids.add(dispimg_id)
                    zip_img_path = dispimg_map[dispimg_id]
                    try:
                        img_bytes = zf.read(zip_img_path)
                        ext = Path(zip_img_path).suffix or ".jpeg"
                        slot_index = scan_col - scan_cfg["start_col"] + 1
                        fname = f"image_{slot_index}{ext}"
                        dest = out_dir / fname
                        dest.write_bytes(img_bytes)
                        image_slots.append(str(dest))
                        logger.info(f"Row {row_idx}, col {get_column_letter(scan_col)}: {dispimg_id} → {dest}")
                    except Exception as e:
                        image_slots.append("")
                        logger.warning(f"Row {row_idx}, col {get_column_letter(scan_col)}: failed to read {zip_img_path}: {e}")
                else:
                    image_slots.append("")
                    empty_streak += 1
                scan_col += 1

        # Fallback: try traditional ws._images if no DISPIMG found
        image_paths = image_slots
        if not any(image_paths):
            for img in ws._images:
                anchor = img.anchor
                img_row = getattr(anchor._from, "row", -1) + 1 if getattr(anchor, "_from", None) else -1
                if img_row == row_idx:
                    ext = Path(img.path).suffix if hasattr(img, "path") and img.path else ".png"
                    slot_index = getattr(anchor._from, "col", len(image_paths)) + 1 if getattr(anchor, "_from", None) else len(image_paths) + 1
                    fname = f"image_{slot_index}{ext}"
                    dest = out_dir / fname
                    with open(dest, "wb") as f:
                        f.write(img._data())
                    while len(image_paths) < slot_index:
                        image_paths.append("")
                    image_paths[slot_index - 1] = str(dest)
                    logger.info(f"Row {row_idx}: fallback ws._images → {dest}")

        while image_paths and not image_paths[-1]:
            image_paths.pop()

        if not any(image_paths):
            logger.warning(f"Row {row_idx} ({product_id}): no images found, skipping")
            skipped += 1
            continue

        products.append(
            ProductRow(
                id=product_id,
                name_cn=name_cn,
                language=language,
                selling_points=selling_points,
                image_paths=image_paths,
                reference_images_are_product=reference_images_are_product,
            )
        )
        if limit is not None and len(products) >= limit:
            break

    wb.close()
    logger.info(f"Parsed {len(products)} products ({skipped} skipped)")
    return products
